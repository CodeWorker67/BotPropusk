import asyncio
import openpyxl
from io import BytesIO
from aiogram import Router, F
from aiogram.filters import CommandStart, Command
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, \
    KeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from sqlalchemy import select, delete

from bot import bot
from config import ADMIN_IDS, RAZRAB
from db.models import Manager, Security, Resident, Contractor, RegistrationRequest, \
    ContractorRegistrationRequest, AsyncSessionLocal, ResidentContractorRequest, PermanentPass, TemporaryPass, Appeal
from filters import IsAdminOrManager

router = Router()
router.message.filter(IsAdminOrManager())
router.callback_query.filter(IsAdminOrManager())


class AddUserStates(StatesGroup):
    WAITING_PHONE = State()
    CHOOSE_TYPE = State()


class ExportStates(StatesGroup):
    WAITING_FILE = State()


@router.callback_query(F.data == "back_to_main")
async def back_to_main_menu(callback: CallbackQuery):
    await callback.message.edit_text(
        text="Добро пожаловать в Главное меню",
        reply_markup=get_admin_menu()
    )


admin_reply_keyboard = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="Главное меню")]],
    resize_keyboard=True,
    is_persistent=True
)


def is_valid_phone(phone: str) -> bool:
    return len(phone) == 11 and phone.isdigit() and phone[0] == '8'


# Обновленное главное меню админа
def get_admin_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="👥Управление пользователями", callback_data="user_management")],
        [InlineKeyboardButton(text="📝 Регистрация", callback_data="registration_menu")],
        [InlineKeyboardButton(text="🚪 Пропуска", callback_data="passes_menu")],
        [InlineKeyboardButton(text="🔍 Поиск пропуска", callback_data="search_pass")],
        [InlineKeyboardButton(text="📈Статистика", callback_data="statistics_menu")],
        [InlineKeyboardButton(text="📨 Обращения в УК", callback_data="appeals_management")],
        [InlineKeyboardButton(text="📩 Выполнить рассылку", callback_data="posting")],
    ])


def get_admin_user_management_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Менеджеры", callback_data="managers_manage")],
        [InlineKeyboardButton(text="СБ", callback_data="security_manage")],
        [InlineKeyboardButton(text="Резиденты", callback_data="residents_manage")],
        [InlineKeyboardButton(text="Подрядчики", callback_data="contractors_manage")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_main")]
        ])


def get_manager_user_management_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Резиденты", callback_data="residents_manage")],
        [InlineKeyboardButton(text="Подрядчики", callback_data="contractors_manage")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_main")]
        ])


def get_add_menu(user_type: str):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"Добавить {user_type}", callback_data=f"add_{user_type}")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_manage")]
        ])


@router.message(CommandStart())
async def process_start_admin(message: Message):
    try:
        await message.answer(
            text="Здравствуйте",
            reply_markup=admin_reply_keyboard
        )
        await message.answer(
            text="Добро пожаловать в Главное меню",
            reply_markup=get_admin_menu()
        )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{message.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message(F.text == "Главное меню")
async def main_menu(message: Message, state: FSMContext):
    try:
        await state.clear()
        await message.answer(
            text="Добро пожаловать в Главное меню",
            reply_markup=get_admin_menu()
        )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{message.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.in_({"user_management", "back_to_manage"}))
async def user_management(callback: CallbackQuery):
    try:
        if callback.from_user.id in ADMIN_IDS:
            kb = get_admin_user_management_menu()
        else:
            kb = get_manager_user_management_menu()
        await callback.message.edit_text(
            text="Выберите категорию пользователей:",
            reply_markup=kb
            )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.endswith("manage"))
async def manage_category(callback: CallbackQuery, state: FSMContext):
    try:
        user_type = callback.data.split("_")[0]
        await state.update_data(user_type=user_type)

        # Для резидентов
        if user_type == 'residents':
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Добавить резидента", callback_data=f"add_{user_type}")],
                [InlineKeyboardButton(text="Список резидентов", callback_data="list_residents")],
                [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_manage")]
            ])
        # Для подрядчиков
        elif user_type == 'contractors':
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Добавить подрядчика", callback_data=f"add_{user_type}")],
                [InlineKeyboardButton(text="Список подрядчиков", callback_data="list_contractors")],
                [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_manage")]
            ])
        elif user_type == 'managers':
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Добавить менеджера", callback_data=f"add_{user_type}")],
                [InlineKeyboardButton(text="Список менеджеров", callback_data="list_managers")],
                [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_manage")]
            ])
            # Для СБ
        elif user_type == 'security':
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Добавить СБ", callback_data=f"add_{user_type}")],
                [InlineKeyboardButton(text="Список СБ", callback_data="list_security")],
                [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_manage")]
            ])
        else:
            return

        await callback.message.edit_text(
            text=f"Управление {user_type}:",
            reply_markup=keyboard
        )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.startswith("add_"))
async def start_add_user(callback: CallbackQuery, state: FSMContext):
    try:
        await callback.message.answer("Введите телефон пользователя:")
        await state.set_state(AddUserStates.WAITING_PHONE)
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message(F.text, AddUserStates.WAITING_PHONE)
async def process_phone(message: Message, state: FSMContext):
    try:
        data = await state.get_data()
        user_type = data['user_type']
        phone = message.text
        if not is_valid_phone(phone):
            await message.answer('Телефон должен быть в формате 8XXXXXXXXXX.\nПопробуйте ввести еще раз!')
            return

        async with AsyncSessionLocal() as session:
            try:
                if user_type == 'managers':
                    new_user = Manager(phone=phone)
                elif user_type == 'security':
                    new_user = Security(phone=phone)
                elif user_type == 'residents':
                    new_user = Resident(phone=phone)
                elif user_type == 'contractors':
                    new_user = Contractor(phone=phone)

                session.add(new_user)
                await session.commit()
                await message.answer(f"Пользователь с телефоном {phone} добавлен в {user_type}!")
                if user_type == 'residents':
                    keyboard = InlineKeyboardMarkup(inline_keyboard=[
                        [InlineKeyboardButton(text="Добавить резидента", callback_data=f"add_{user_type}")],
                        [InlineKeyboardButton(text="Список резидентов", callback_data="list_residents")],
                        [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_manage")]
                    ])
                # Для подрядчиков
                elif user_type == 'contractors':
                    keyboard = InlineKeyboardMarkup(inline_keyboard=[
                        [InlineKeyboardButton(text="Добавить подрядчика", callback_data=f"add_{user_type}")],
                        [InlineKeyboardButton(text="Список подрядчиков", callback_data="list_contractors")],
                        [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_manage")]
                    ])
                # Для остальных
                elif user_type == 'managers':
                    keyboard = InlineKeyboardMarkup(inline_keyboard=[
                        [InlineKeyboardButton(text="Добавить менеджера", callback_data=f"add_{user_type}")],
                        [InlineKeyboardButton(text="Список менеджеров", callback_data="list_managers")],
                        [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_manage")]
                    ])
                    # Для СБ
                elif user_type == 'security':
                    keyboard = InlineKeyboardMarkup(inline_keyboard=[
                        [InlineKeyboardButton(text="Добавить СБ", callback_data=f"add_{user_type}")],
                        [InlineKeyboardButton(text="Список СБ", callback_data="list_security")],
                        [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_manage")]
                    ])
                else:
                    return

                await message.answer(
                    text=f"Управление {user_type}:",
                    reply_markup=keyboard
                )

            except Exception as e:
                await message.answer(f"Ошибка: {str(e)}")
                await session.rollback()

    except Exception as e:
        await bot.send_message(RAZRAB, f'{message.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data == "list_residents")
async def show_residents_list(callback: CallbackQuery):
    try:
        async with AsyncSessionLocal() as session:
            # Получаем всех резидентов со статусом True
            result = await session.execute(
                select(Resident).where(Resident.status == True))
            residents = result.scalars().all()

            if not residents:
                await callback.answer("Нет зарегистрированных резидентов")
                return

            buttons = []
            for resident in residents:
                # Формируем текст кнопки: ID и ФИО
                button_text = f"{resident.fio}"
                # Укорачиваем, если слишком длинное
                if len(button_text) > 30:
                    button_text = button_text[:27] + "..."

                buttons.append([InlineKeyboardButton(
                    text=button_text,
                    callback_data=f"view_resident_{resident.id}"
                )])

            # Добавляем кнопку "Назад"
            buttons.append([InlineKeyboardButton(
                text="⬅️ Назад",
                callback_data="residents_manage"
            )])
            try:
                await callback.message.edit_text(
                    "Список зарегистрированных резидентов:",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
                )
            except:
                await callback.message.answer(text="Список зарегистрированных резидентов:",
                                              reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.startswith("view_resident_"))
async def view_resident_details(callback: CallbackQuery):
    try:
        resident_id = int(callback.data.split("_")[-1])
        async with AsyncSessionLocal() as session:
            resident = await session.get(Resident, resident_id)
            if not resident:
                await callback.answer("Резидент не найден")
                return

            # Формируем текст
            text = (
                f"ID: {resident.id}\n"
                f"ФИО: {resident.fio}\n"
                f"Телефон: {resident.phone}\n"
                f"Номер участка: {resident.plot_number}\n"
                f"TG: @{resident.username}\n"
                f"Время регистрации: {resident.time_registration}"
            )

            # Клавиатура с кнопкой "Назад" к списку резидентов
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🗑 Удалить", callback_data=f"delete_resident_{resident_id}")],
                [InlineKeyboardButton(text="⬅️ Назад", callback_data="list_residents")]
            ])

            await callback.message.edit_text(
                text=text,
                reply_markup=keyboard
            )
            await callback.answer()
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data == "list_contractors")
async def show_contractors_list(callback: CallbackQuery):
    try:
        async with AsyncSessionLocal() as session:
            # Получаем всех подрядчиков со статусом True
            result = await session.execute(
                select(Contractor).where(Contractor.status == True))
            contractors = result.scalars().all()

            if not contractors:
                await callback.answer("Нет зарегистрированных подрядчиков")
                return

            buttons = []
            for contractor in contractors:
                # Формируем текст кнопки: ID и ФИО
                button_text = f"{contractor.company}_{contractor.position}"
                # Укорачиваем, если слишком длинное
                if len(button_text) > 30:
                    button_text = button_text[:27] + "..."

                buttons.append([InlineKeyboardButton(
                    text=button_text,
                    callback_data=f"view_contractor_{contractor.id}"
                )])

            # Добавляем кнопку "Назад"
            buttons.append([InlineKeyboardButton(
                text="⬅️ Назад",
                callback_data="contractors_manage"
            )])

            try:
                await callback.message.edit_text(
                    "Список зарегистрированных подрядчиков:",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
                )
            except:
                await callback.message.answer(
                    text="Список зарегистрированных подрядчиков:",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
                )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.startswith("view_contractor_"))
async def view_contractor_details(callback: CallbackQuery):
    try:
        contractor_id = int(callback.data.split("_")[-1])
        async with AsyncSessionLocal() as session:
            contractor = await session.get(Contractor, contractor_id)
            if not contractor:
                await callback.answer("Подрядчик не найден")
                return

            # Формируем текст
            text = (
                f"ID: {contractor.id}\n"
                f"ФИО: {contractor.fio}\n"
                f"Телефон: {contractor.phone}\n"
                f"Компания: {contractor.company}\n"
                f"Должность: {contractor.position}\n"
                f"Принадлежность: {contractor.affiliation}\n"
                f"TG: @{contractor.username}\n"
                f"Возможность добавлять субподрядчиков: {contractor.can_add_contractor}\n"
                f"Время регистрации: {contractor.time_registration}"
            )
            if contractor.can_add_contractor == True:
                text_admin = '✅Подрядчик-администратор'
            else:
                text_admin = '❌Подрядчик-администратор'

            # Клавиатура с кнопкой "Назад" к списку подрядчиков
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🗑 Удалить", callback_data=f"delete_contractor_{contractor_id}")],
                [InlineKeyboardButton(text=text_admin, callback_data=f"change_admin_{contractor_id}")],
                [InlineKeyboardButton(text="⬅️ Назад", callback_data="list_contractors")]
            ])

            await callback.message.edit_text(
                text=text,
                reply_markup=keyboard
            )
            await callback.answer()
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.startswith("delete_resident_"))
async def confirm_delete_resident(callback: CallbackQuery):
    try:
        resident_id = int(callback.data.split("_")[-1])
        await callback.message.answer(
            "Вы точно хотите удалить резидента?",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✅ Да", callback_data=f"confirm_delete_yes_{resident_id}")],
                [InlineKeyboardButton(text="❌ Нет", callback_data=f"confirm_delete_no_{resident_id}")]
            ])
        )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.startswith("confirm_delete_no_"))
async def cancel_delete(callback: CallbackQuery):
    try:
        resident_id = int(callback.data.split("_")[-1])
        # Возвращаемся к просмотру резидента
        await view_resident_details(callback)
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.startswith("confirm_delete_yes_"))
async def execute_delete(callback: CallbackQuery, state: FSMContext):
    try:
        resident_id = int(callback.data.split("_")[-1])
        async with AsyncSessionLocal() as session:
            # Удаляем связанные заявки
            stmt1 = delete(RegistrationRequest).where(RegistrationRequest.resident_id == resident_id)
            stmt2 = delete(ResidentContractorRequest).where(ResidentContractorRequest.resident_id == resident_id)
            stmt3 = delete(PermanentPass).where(PermanentPass.resident_id == resident_id)
            stmt4 = delete(TemporaryPass).where(TemporaryPass.resident_id == resident_id)
            stmt5 = delete(Appeal).where(Appeal.resident_id == resident_id)
            await session.execute(stmt1)
            await session.execute(stmt2)
            await session.execute(stmt3)
            await session.execute(stmt4)
            await session.execute(stmt5)
            resident = await session.get(Resident, resident_id)
            await bot.send_message(resident.tg_id, 'Вам ограничили доступ, если это случилось по ошибке обратитесь в управляющую компанию "Ели Estate"')
            # Удаляем резидента
            stmt6 = delete(Resident).where(Resident.id == resident_id)
            await session.execute(stmt6)
            await session.commit()

        await callback.message.answer("✅ Резидент удален")
        # Возвращаемся в меню управления резидентами
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Добавить резидента", callback_data=f"add_residents")],
            [InlineKeyboardButton(text="Список резидентов", callback_data="list_residents")],
            [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_manage")]
        ])
        await callback.message.answer(
            text=f"Управление residents:",
            reply_markup=keyboard
        )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


# Подтверждение удаления подрядчика
@router.callback_query(F.data.startswith("delete_contractor_"))
async def confirm_delete_contractor(callback: CallbackQuery):
    try:
        contractor_id = int(callback.data.split("_")[-1])
        await callback.message.answer(
            "Вы точно хотите удалить подрядчика?",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✅ Да", callback_data=f"confirm_del_cont_yes_{contractor_id}")],
                [InlineKeyboardButton(text="❌ Нет", callback_data=f"confirm_del_cont_no_{contractor_id}")]
            ])
        )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


# Отмена удаления подрядчика
@router.callback_query(F.data.startswith("confirm_del_cont_no_"))
async def cancel_delete_contractor(callback: CallbackQuery):
    try:
        contractor_id = int(callback.data.split("_")[-1])
        # Возвращаемся к просмотру подрядчика
        await view_contractor_details(callback)
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


# Выполнение удаления подрядчика
@router.callback_query(F.data.startswith("confirm_del_cont_yes_"))
async def execute_delete_contractor(callback: CallbackQuery, state: FSMContext):
    try:
        contractor_id = int(callback.data.split("_")[-1])
        async with AsyncSessionLocal() as session:
            # Удаляем связанные записи
            stmt1 = delete(ContractorRegistrationRequest).where(
                ContractorRegistrationRequest.contractor_id == contractor_id
            )
            stmt2 = delete(TemporaryPass).where(
                TemporaryPass.contractor_id == contractor_id
            )
            await session.execute(stmt1)
            await session.execute(stmt2)
            contractor = await session.get(Contractor, contractor_id)
            await bot.send_message(contractor.tg_id,
                                   'Вам ограничили доступ, если это случилось по ошибке обратитесь в управляющую компанию "Ели Estate"')
            # Удаляем самого подрядчика
            stmt3 = delete(Contractor).where(Contractor.id == contractor_id)
            await session.execute(stmt3)
            await session.commit()

        await callback.message.answer("✅ Подрядчик удален")

        # Возвращаемся к списку подрядчиков
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Добавить подрядчика", callback_data=f"add_contractors")],
            [InlineKeyboardButton(text="Список подрядчиков", callback_data="list_contractors")],
            [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_manage")]
        ])
        await callback.message.answer(
            text=f"Управление contractors:",
            reply_markup=keyboard
        )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data == "list_managers")
async def show_managers_list(callback: CallbackQuery):
    try:
        async with AsyncSessionLocal() as session:
            result = await session.execute(
                select(Manager).where(Manager.status == True))
            managers = result.scalars().all()

            if not managers:
                await callback.answer("Нет зарегистрированных менеджеров")
                return

            buttons = []
            for manager in managers:
                button_text = f"{manager.fio}"
                if len(button_text) > 30:
                    button_text = button_text[:27] + "..."

                buttons.append([InlineKeyboardButton(
                    text=button_text,
                    callback_data=f"view_manager_{manager.id}"
                )])

            buttons.append([InlineKeyboardButton(
                text="⬅️ Назад",
                callback_data="managers_manage"
            )])

            await callback.message.edit_text(
                "Список менеджеров:",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
            )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data == "list_security")
async def show_security_list(callback: CallbackQuery):
    try:
        async with AsyncSessionLocal() as session:
            result = await session.execute(
                select(Security).where(Security.status == True))
            security_list = result.scalars().all()

            if not security_list:
                await callback.answer("Нет зарегистрированных сотрудников СБ")
                return

            buttons = []
            for security in security_list:
                button_text = f"{security.fio}"
                if len(button_text) > 30:
                    button_text = button_text[:27] + "..."

                buttons.append([InlineKeyboardButton(
                    text=button_text,
                    callback_data=f"view_security_{security.id}"
                )])

            buttons.append([InlineKeyboardButton(
                text="⬅️ Назад",
                callback_data="security_manage"
            )])

            await callback.message.edit_text(
                "Список сотрудников СБ:",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
            )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.startswith("view_manager_"))
async def view_manager_details(callback: CallbackQuery):
    try:
        manager_id = int(callback.data.split("_")[-1])
        async with AsyncSessionLocal() as session:
            manager = await session.get(Manager, manager_id)
            if not manager:
                await callback.answer("Менеджер не найден")
                return

            text = (
                f"ID: {manager.id}\n"
                f"ФИО: {manager.fio}\n"
                f"Телефон: {manager.phone}\n"
                f"Username: @{manager.username}\n"
                f"TG ID: {manager.tg_id}\n"
                f"Время добавления: {manager.time_add_to_db}\n"
                f"Время регистрации: {manager.time_registration}\n"
                f"Статус: {'Активен' if manager.status else 'Неактивен'}"
            )

            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🗑 Удалить", callback_data=f"delete_manager_{manager_id}")],
                [InlineKeyboardButton(text="⬅️ Назад", callback_data="list_managers")]
            ])

            await callback.message.edit_text(
                text=text,
                reply_markup=keyboard
            )
            await callback.answer()
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)

@router.callback_query(F.data.startswith("view_security_"))
async def view_security_details(callback: CallbackQuery):
    try:
        security_id = int(callback.data.split("_")[-1])
        async with AsyncSessionLocal() as session:
            security = await session.get(Security, security_id)
            if not security:
                await callback.answer("Сотрудник СБ не найден")
                return

            text = (
                f"ID: {security.id}\n"
                f"ФИО: {security.fio}\n"
                f"Телефон: {security.phone}\n"
                f"Username: @{security.username}\n"
                f"TG ID: {security.tg_id}\n"
                f"Время добавления: {security.time_add_to_db}\n"
                f"Время регистрации: {security.time_registration}\n"
                f"Статус: {'Активен' if security.status else 'Неактивен'}"
            )

            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🗑 Удалить", callback_data=f"delete_security_{security_id}")],
                [InlineKeyboardButton(text="⬅️ Назад", callback_data="list_security")]
            ])

            await callback.message.edit_text(
                text=text,
                reply_markup=keyboard
            )
            await callback.answer()
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.startswith("delete_manager_"))
async def confirm_delete_manager(callback: CallbackQuery):
    try:
        manager_id = int(callback.data.split("_")[-1])
        await callback.message.edit_text(
            "Вы точно хотите удалить менеджера?",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✅ Да", callback_data=f"confirm_delete_manager_yes_{manager_id}")],
                [InlineKeyboardButton(text="❌ Нет", callback_data=f"confirm_delete_manager_no_{manager_id}")]
            ])
        )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)

@router.callback_query(F.data.startswith("delete_security_"))
async def confirm_delete_security(callback: CallbackQuery):
    try:
        security_id = int(callback.data.split("_")[-1])
        await callback.message.edit_text(
            "Вы точно хотите удалить сотрудника СБ?",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✅ Да", callback_data=f"confirm_delete_security_yes_{security_id}")],
                [InlineKeyboardButton(text="❌ Нет", callback_data=f"confirm_delete_security_no_{security_id}")]
            ])
        )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.startswith("confirm_delete_manager_yes_"))
async def execute_delete_manager(callback: CallbackQuery, state: FSMContext):
    try:
        manager_id = int(callback.data.split("_")[-1])
        async with AsyncSessionLocal() as session:
            manager = await session.get(Manager, manager_id)
            await bot.send_message(manager.tg_id,
                                   'Вам ограничили доступ, если это случилось по ошибке обратитесь в управляющую компанию "Ели Estate"')
            stmt = delete(Manager).where(Manager.id == manager_id)
            await session.execute(stmt)
            await session.commit()

        await callback.message.answer("✅ Менеджер удален")

        # Возвращаемся в меню управления менеджерами
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Добавить менеджера", callback_data=f"add_managers")],
            [InlineKeyboardButton(text="Список менеджеров", callback_data="list_managers")],
            [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_manage")]
        ])

        await callback.message.answer(
            text=f"Управление managers:",
            reply_markup=keyboard
        )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.startswith("confirm_delete_security_yes_"))
async def execute_delete_security(callback: CallbackQuery, state: FSMContext):
    try:
        security_id = int(callback.data.split("_")[-1])
        async with AsyncSessionLocal() as session:
            security = await session.get(Security, security_id)
            await bot.send_message(security.tg_id,
                                   'Вам ограничили доступ, если это случилось по ошибке обратитесь в управляющую компанию "Ели Estate"')
            stmt = delete(Security).where(Security.id == security_id)
            await session.execute(stmt)
            await session.commit()

        await callback.message.answer("✅ Сотрудник СБ удален")

        # Возвращаемся в меню управления СБ
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Добавить СБ", callback_data=f"add_security")],
            [InlineKeyboardButton(text="Список СБ", callback_data="list_security")],
            [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_manage")]
        ])

        await callback.message.answer(
            text=f"Управление security",
            reply_markup=keyboard
        )
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.startswith("confirm_delete_manager_no_"))
async def execute_no_delete_manager(callback: CallbackQuery, state: FSMContext):
    try:
        manager_id = int(callback.data.split("_")[-1])
        async with AsyncSessionLocal() as session:
            manager = await session.get(Manager, manager_id)
            if not manager:
                await callback.answer("Менеджер не найден")
                return

            text = (
                f"ID: {manager.id}\n"
                f"ФИО: {manager.fio}\n"
                f"Телефон: {manager.phone}\n"
                f"Username: @{manager.username}\n"
                f"TG ID: {manager.tg_id}\n"
                f"Время добавления: {manager.time_add_to_db}\n"
                f"Время регистрации: {manager.time_registration}\n"
                f"Статус: {'Активен' if manager.status else 'Неактивен'}"
            )

            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🗑 Удалить", callback_data=f"delete_manager_{manager_id}")],
                [InlineKeyboardButton(text="⬅️ Назад", callback_data="list_managers")]
            ])

            await callback.message.edit_text(
                text=text,
                reply_markup=keyboard
            )
            await callback.answer()
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.startswith("confirm_delete_security_no_"))
async def execute_no_delete_security(callback: CallbackQuery, state: FSMContext):
    try:
        security_id = int(callback.data.split("_")[-1])
        async with AsyncSessionLocal() as session:
            security = await session.get(Security, security_id)
            if not security:
                await callback.answer("Сотрудник СБ не найден")
                return

            text = (
                f"ID: {security.id}\n"
                f"ФИО: {security.fio}\n"
                f"Телефон: {security.phone}\n"
                f"Username: @{security.username}\n"
                f"TG ID: {security.tg_id}\n"
                f"Время добавления: {security.time_add_to_db}\n"
                f"Время регистрации: {security.time_registration}\n"
                f"Статус: {'Активен' if security.status else 'Неактивен'}"
            )

            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🗑 Удалить", callback_data=f"delete_security_{security_id}")],
                [InlineKeyboardButton(text="⬅️ Назад", callback_data="list_security")]
            ])

            await callback.message.edit_text(
                text=text,
                reply_markup=keyboard
            )
            await callback.answer()
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.callback_query(F.data.startswith("change_admin_"))
async def change_contractor_admin(callback: CallbackQuery):
    try:
        contractor_id = int(callback.data.split("_")[-1])
        async with AsyncSessionLocal() as session:
            contractor = await session.get(Contractor, contractor_id)
            if not contractor:
                await callback.answer("Подрядчик не найден")
                return

            # Формируем текст
            text = (
                f"ID: {contractor.id}\n"
                f"ФИО: {contractor.fio}\n"
                f"Телефон: {contractor.phone}\n"
                f"Компания: {contractor.company}\n"
                f"Должность: {contractor.position}\n"
                f"Принадлежность: {contractor.affiliation}\n"
                f"TG: @{contractor.username}\n"
                f"Возможность добавлять субподрядчиков: {not contractor.can_add_contractor}\n"
                f"Время регистрации: {contractor.time_registration}"
            )
            if contractor.can_add_contractor == False:
                text_admin = '✅Подрядчик-администратор'
            else:
                text_admin = '❌Подрядчик-администратор'

            # Клавиатура с кнопкой "Назад" к списку подрядчиков
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🗑 Удалить", callback_data=f"delete_contractor_{contractor_id}")],
                [InlineKeyboardButton(text=text_admin, callback_data=f"change_admin_{contractor_id}")],
                [InlineKeyboardButton(text="⬅️ Назад", callback_data="list_contractors")]
            ])
            await callback.message.edit_text(
                text=text,
                reply_markup=keyboard
            )
            await callback.answer()
            contractor.can_add_contractor = not contractor.can_add_contractor
            await session.commit()
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message(Command("import"))
async def command_export(message: Message, state: FSMContext):
    try:
        await state.clear()
        await message.answer("Пожалуйста, загрузите Excel-файл с резидентами.")
        await state.set_state(ExportStates.WAITING_FILE)
    except Exception as e:
        await bot.send_message(RAZRAB, f'{message.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message(ExportStates.WAITING_FILE, F.document)
async def handle_export_file(message: Message, state: FSMContext):
    excel_file = None
    try:
        if not message.document.file_name.endswith('.xlsx'):
            await message.answer("Пожалуйста, загрузите файл в формате xlsx")
            return

        file_id = message.document.file_id
        file = await bot.get_file(file_id)
        file_path = file.file_path

        excel_file = BytesIO()
        await bot.download_file(file_path, excel_file)
        excel_file.seek(0)

        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        errors = []
        success_count = 0
        total_rows = sheet.max_row

        async with AsyncSessionLocal() as session:
            for row in range(1, total_rows + 1):
                phone_cell = sheet.cell(row=row, column=1)
                fio_cell = sheet.cell(row=row, column=2)
                phone = str(phone_cell.value) if phone_cell.value is not None else ""
                fio = fio_cell.value

                if not phone or not is_valid_phone(phone) or not fio:
                    errors.append(f"Строка №{row} - не корректный телефон или фио должно быть заполнено")
                    continue

                try:
                    resident = Resident(phone=phone, fio=fio)
                    session.add(resident)
                    success_count += 1
                except Exception as e:
                    errors.append(f"Строка №{row} - ошибка при добавлении в базу: {str(e)}")

            await session.commit()

        report = f"Загружено {success_count} резидентов из {total_rows} строк."
        if errors:
            error_report = "\n".join(errors)
            report += f"\nОшибки:\n{error_report}"

        if len(report) > 4096:
            for x in range(0, len(report), 4096):
                await message.answer(report[x:x+4096])
        else:
            await message.answer(report)

    except Exception as e:
        await bot.send_message(RAZRAB, f'{message.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)
    finally:
        if excel_file:
            excel_file.close()
        await state.clear()
