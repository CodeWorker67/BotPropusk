# handlers_admin_statistic.py
import asyncio
import logging
from datetime import date, datetime
from io import BytesIO

import openpyxl
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile
from openpyxl.workbook import Workbook
from sqlalchemy import and_, func, select

from bot import bot
from db.models import (
    AsyncSessionLocal,
    Resident,
    Contractor,
    PermanentPass,
    TemporaryPass,
    TempPassYooKassaPayment,
    Manager,
    Security,
)
from config import ADMIN_IDS, RAZRAB
from filters import IsAdminOrManager
from temporary_truck import temp_pass_last_valid_date

router = Router()
router.message.filter(IsAdminOrManager())
router.callback_query.filter(IsAdminOrManager())


@router.callback_query(F.data == "statistics_menu")
async def show_statistics(callback: CallbackQuery):
    try:
        async with AsyncSessionLocal() as session:
            # Статистика по резидентам
            total_residents = await session.scalar(select(func.count(Resident.id)))
            registered_residents = await session.scalar(
                select(func.count(Resident.id))
                .where(Resident.status == True)
            )
            unregistered_residents = total_residents - registered_residents

            # Статистика по подрядчикам
            total_contractors = await session.scalar(select(func.count(Contractor.id)))
            registered_contractors = await session.scalar(
                select(func.count(Contractor.id))
                .where(Contractor.status == True)
            )
            unregistered_contractors = total_contractors - registered_contractors

            # Статистика по постоянным пропускам
            total_permanent = await session.scalar(select(func.count(PermanentPass.id)))
            pending_permanent = await session.scalar(
                select(func.count(PermanentPass.id))
                .where(PermanentPass.status == 'pending')
            )
            approved_permanent = await session.scalar(
                select(func.count(PermanentPass.id))
                .where(PermanentPass.status == 'approved')
            )
            rejected_permanent = await session.scalar(
                select(func.count(PermanentPass.id))
                .where(PermanentPass.status == 'rejected')
            )

            # Статистика по временным пропускам
            total_temporary = await session.scalar(select(func.count(TemporaryPass.id)))
            pending_temporary = await session.scalar(
                select(func.count(TemporaryPass.id))
                .where(TemporaryPass.status == 'pending')
            )
            approved_temporary = await session.scalar(
                select(func.count(TemporaryPass.id))
                .where(TemporaryPass.status == 'approved')
            )
            rejected_temporary = await session.scalar(
                select(func.count(TemporaryPass.id))
                .where(TemporaryPass.status == 'rejected')
            )

            total_passes = total_permanent + total_temporary
            pending_passes = pending_permanent + pending_temporary
            approved_passes = approved_permanent + approved_temporary
            rejected_passes = rejected_temporary + rejected_permanent

        # Формируем сообщение
        text = (
            "📊 <b>Статистика системы</b>\n\n"
            "👤 <b>Резиденты:</b>\n"
            f"  Всего: {total_residents}\n"
            f"  Зарегистрированных: {registered_residents}\n"
            f"  Не зарегистрированных: {unregistered_residents}\n\n"
    
            "👷 <b>Подрядчики:</b>\n"
            f"  Всего: {total_contractors}\n"
            f"  Зарегистрированных: {registered_contractors}\n"
            f"  Не зарегистрированных: {unregistered_contractors}\n\n"
    
            "🎫 <b>Все пропуска:</b>\n"
            f"  Всего заявок: {total_passes}\n"
            f"  На утверждении: {pending_passes}\n"
            f"  Утвержденных: {approved_passes}\n"
            f"  Отклоненных: {rejected_passes}\n\n"
    
            "🔖 <b>Постоянные пропуска:</b>\n"
            f"  Всего заявок: {total_permanent}\n"
            f"  На утверждении: {pending_permanent}\n"
            f"  Утвержденных: {approved_permanent}\n"
            f"  Отклоненных: {rejected_permanent}\n\n"
    
            "⏳ <b>Временные пропуска:</b>\n"
            f"  Всего заявок: {total_temporary}\n"
            f"  На утверждении: {pending_temporary}\n"
            f"  Утвержденных: {approved_temporary}\n"
            f"  Отклоненных: {rejected_temporary}"
        )

        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📤 Экспорт в xlsx", callback_data="export_to_xlsx")],
            [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_main")]
        ])
        try:
            await callback.message.edit_text(text, reply_markup=keyboard, parse_mode="HTML")
        except:
            await callback.message.answer(text, reply_markup=keyboard, parse_mode="HTML")
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)


# Новый обработчик для экспорта
@router.callback_query(F.data == "export_to_xlsx")
async def export_statistics_to_xlsx(callback: CallbackQuery):
    try:
        await callback.answer("Формируем отчет...")
        async with AsyncSessionLocal() as session:
            wb = Workbook()

            # Лист: Резиденты
            ws_res = wb.active
            ws_res.title = "Резиденты"
            residents = await session.execute(select(Resident))
            ws_res.append(["ID", "Телефон", "ФИО", "Участок", "TG ID", "Статус"])
            for res in residents.scalars():
                ws_res.append([
                    res.id, res.phone, res.fio, res.plot_number,
                    res.tg_id, "Активен" if res.status else "Неактивен"
                ])

            # Лист: Подрядчики
            ws_contr = wb.create_sheet("Подрядчики")
            contractors = await session.execute(select(Contractor))
            ws_contr.append(["ID", "Телефон", "ФИО", "Компания", "Должность", "TG ID", "Статус"])
            for contr in contractors.scalars():
                ws_contr.append([
                    contr.id, contr.phone, contr.fio, contr.company,
                    contr.position, contr.tg_id, "Активен" if contr.status else "Неактивен"
                ])

            def _dt(v):
                return v.strftime("%Y-%m-%d %H:%M:%S") if v else ""

            # Лист: Менеджеры
            ws_mgr = wb.create_sheet("Менеджеры")
            managers = await session.execute(select(Manager))
            ws_mgr.append([
                "ID", "Стартовый ключ", "Телефон", "ФИО", "TG ID", "Username",
                "Имя TG", "Фамилия TG", "Добавлен в БД", "Регистрация", "Статус",
            ])
            for mgr in managers.scalars():
                ws_mgr.append([
                    mgr.id,
                    mgr.start_key,
                    mgr.phone,
                    mgr.fio,
                    mgr.tg_id,
                    mgr.username,
                    mgr.first_name,
                    mgr.last_name,
                    _dt(mgr.time_add_to_db),
                    _dt(mgr.time_registration),
                    "Активен" if mgr.status else "Неактивен",
                ])

            # Лист: СБ
            ws_sec = wb.create_sheet("СБ")
            securities = await session.execute(select(Security))
            ws_sec.append([
                "ID", "Стартовый ключ", "Телефон", "ФИО", "TG ID", "Username",
                "Имя TG", "Фамилия TG", "Добавлен в БД", "Регистрация", "Статус",
            ])
            for sec in securities.scalars():
                ws_sec.append([
                    sec.id,
                    sec.start_key,
                    sec.phone,
                    sec.fio,
                    sec.tg_id,
                    sec.username,
                    sec.first_name,
                    sec.last_name,
                    _dt(sec.time_add_to_db),
                    _dt(sec.time_registration),
                    "Активен" if sec.status else "Неактивен",
                ])

            # Лист: Постоянные пропуска
            ws_perm = wb.create_sheet("Постоянные пропуска")
            stmt = select(
                PermanentPass,
                Resident.fio,
                Resident.plot_number
            ).join(Resident, PermanentPass.resident_id == Resident.id)

            passes = await session.execute(stmt)
            ws_perm.append([
                "ID", "Резидент ID", "ФИО резидента", "Участок",
                "Марка", "Модель", "Номер", "Владелец", "Статус"
            ])

            for pass_data in passes:
                pp = pass_data[0]
                ws_perm.append([
                    pp.id, pp.resident_id, pass_data[1], pass_data[2],
                    pp.car_brand, pp.car_model, pp.car_number,
                    pp.car_owner, pp.status
                ])

            # Лист: Временные пропуска
            ws_temp = wb.create_sheet("Временные пропуска")
            headers = [
                "ID", "Тип владельца", "ФИО", "Участок/Компания", "Должность",
                "Тип ТС", "Категория веса", "Категория длины", "Номер авто",
                "Марка", "Груз", "Цель", "Дата визита", "Статус"
            ]
            ws_temp.append(headers)

            # Отдельные запросы для резидентов и подрядчиков
            # Для пропусков типа "resident"
            res_stmt = select(
                TemporaryPass,
                Resident.fio,
                Resident.plot_number
            ).join(Resident, TemporaryPass.resident_id == Resident.id) \
                .where(TemporaryPass.owner_type == "resident")

            res_temp_passes = await session.execute(res_stmt)
            today = date.today()
            for tp_data in res_temp_passes:
                tp = tp_data[0]
                if temp_pass_last_valid_date(tp.visit_date, tp.purpose) < today:
                    continue
                ws_temp.append([
                    tp.id,
                    "Резидент",
                    tp_data[1],  # fio
                    tp_data[2],  # plot_number
                    "",  # должность отсутствует
                    tp.vehicle_type,
                    tp.weight_category,
                    tp.length_category,
                    tp.car_number,
                    tp.car_brand,
                    tp.cargo_type,
                    tp.purpose,
                    tp.visit_date.strftime("%Y-%m-%d"),
                    tp.status
                ])

            # Для пропусков типа "contractor"
            contr_stmt = select(
                TemporaryPass,
                Contractor.fio,
                Contractor.company,
                Contractor.position
            ).join(Contractor, TemporaryPass.contractor_id == Contractor.id) \
                .where(TemporaryPass.owner_type == "contractor")

            contr_temp_passes = await session.execute(contr_stmt)
            for tp_data in contr_temp_passes:
                tp = tp_data[0]
                if temp_pass_last_valid_date(tp.visit_date, tp.purpose) < today:
                    continue
                ws_temp.append([
                    tp.id,
                    "Подрядчик",
                    tp_data[1],  # fio
                    tp_data[2],  # company
                    tp_data[3],  # position
                    tp.vehicle_type,
                    tp.weight_category,
                    tp.length_category,
                    tp.car_number,
                    tp.car_brand,
                    tp.cargo_type,
                    tp.purpose,
                    tp.visit_date.strftime("%Y-%m-%d"),
                    tp.status
                ])

            # Лист: оплаченные грузовые пропуска (одобрены и оплачены)
            ws_paid_passes = wb.create_sheet("Оплаченные пропуска")
            ws_paid_passes.append([
                "ID пропуска",
                "Тип владельца",
                "ФИО",
                "Участок/Компания",
                "Должность",
                "Категория веса",
                "Номер авто",
                "Марка",
                "Дата визита",
                "Статус пропуска",
                "Сумма оплаты (руб)",
                "Дата оплаты",
            ])
            pay_join = and_(
                TempPassYooKassaPayment.temporary_pass_id == TemporaryPass.id,
                TempPassYooKassaPayment.status == "succeeded",
            )
            truck_approved = and_(
                TemporaryPass.vehicle_type == "truck",
                TemporaryPass.status == "approved",
            )

            paid_pass_rows: list[tuple] = []

            res_paid_stmt = (
                select(
                    TemporaryPass,
                    Resident.fio,
                    Resident.plot_number,
                    TempPassYooKassaPayment.amount_kopeks,
                    TempPassYooKassaPayment.paid_at,
                )
                .join(Resident, TemporaryPass.resident_id == Resident.id)
                .join(TempPassYooKassaPayment, pay_join)
                .where(
                    TemporaryPass.owner_type == "resident",
                    truck_approved,
                )
            )
            for row in await session.execute(res_paid_stmt):
                tp, fio, plot, amount_k, paid_at = row
                paid_pass_rows.append(
                    (
                        paid_at if paid_at is not None else datetime.min,
                        tp,
                        "Резидент",
                        fio,
                        plot,
                        "",
                        amount_k,
                        paid_at,
                    )
                )

            contr_paid_stmt = (
                select(
                    TemporaryPass,
                    Contractor.fio,
                    Contractor.company,
                    Contractor.position,
                    TempPassYooKassaPayment.amount_kopeks,
                    TempPassYooKassaPayment.paid_at,
                )
                .join(Contractor, TemporaryPass.contractor_id == Contractor.id)
                .join(TempPassYooKassaPayment, pay_join)
                .where(
                    TemporaryPass.owner_type == "contractor",
                    truck_approved,
                )
            )
            for row in await session.execute(contr_paid_stmt):
                tp, fio, company, position, amount_k, paid_at = row
                paid_pass_rows.append(
                    (
                        paid_at if paid_at is not None else datetime.min,
                        tp,
                        "Подрядчик",
                        fio,
                        company,
                        position,
                        amount_k,
                        paid_at,
                    )
                )

            paid_pass_rows.sort(key=lambda r: r[0], reverse=True)
            for _, tp, owner_label, fio, plot_or_company, position, amount_k, paid_at in paid_pass_rows:
                ws_paid_passes.append([
                    tp.id,
                    owner_label,
                    fio,
                    plot_or_company,
                    position,
                    tp.weight_category,
                    tp.car_number,
                    tp.car_brand,
                    tp.visit_date.strftime("%Y-%m-%d"),
                    tp.status,
                    amount_k / 100.0,
                    _dt(paid_at),
                ])

            # Сохраняем в буфер
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Отправляем файл
            await callback.message.answer_document(
                document=BufferedInputFile(buffer.read(), filename="Статистика.xlsx"),
                caption="📊 Экспорт статистики завершен"
            )

            # Показываем меню статистики снова
            await show_statistics(callback)
    except Exception as e:
        await bot.send_message(RAZRAB, f'{callback.from_user.id} - {str(e)}')
        await asyncio.sleep(0.05)